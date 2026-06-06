#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Interpreta la estructura de la hoja 'KARDEX VENDEDORAS  2026' para entender:
- cuáles filas son ENCABEZADOS de vendedora (nombre del bloque),
- cuáles son CLIENTES (tienen saldo),
- cuáles son SUBTOTALES (=SUM),
- cuáles tienen #REF! (saldo perdido),
y cuántos clientes hay por vendedora.

Salida a consola: estructura (con nombres de vendedora — dato del dueño, no se commitea).
Salida local: tools/vendedoras.csv (gitignored) con vendedora + nº de clientes.
"""
import csv
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

XLSX = "NUEVO KARDEX KARY DEL 2026....xlsx"
SHEET = "KARDEX VENDEDORAS  2026"
REF = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def detect_saldo_col(ws):
    from collections import Counter
    votes = Counter()
    for row in ws.iter_rows(min_row=2, max_row=min(60, ws.max_row)):
        best_col, best = None, 0
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("=") and ("+" in v or "-" in v):
                n = len(re.findall(r"[A-Z]{1,3}\d+", v))
                if n > best:
                    best, best_col = n, c.column
        if best_col and best >= 4:
            votes[best_col] += 1
    return votes.most_common(1)[0][0] if votes else None


wbF = load_workbook(XLSX, data_only=False, read_only=True)
wbV = load_workbook(XLSX, data_only=True, read_only=True)
wsF, wsV = wbF[SHEET], wbV[SHEET]
scol = detect_saldo_col(wsF)
print(f"Hoja: {SHEET!r} | columna de saldo: {get_column_letter(scol)} (#{scol})\n")

vendedoras = []         # (nombre, n_clientes, n_ref)
cur_name, cur_cli, cur_ref = None, 0, 0
n_sub = 0


def flush():
    global cur_name, cur_cli, cur_ref
    if cur_name is not None:
        vendedoras.append((cur_name, cur_cli, cur_ref))
    cur_name, cur_cli, cur_ref = None, 0, 0


for r in range(2, wsV.max_row + 1):
    a = wsV.cell(r, 1).value
    if a is None or str(a).strip() == "":
        continue
    f = wsF.cell(r, scol).value if scol else None
    v = wsV.cell(r, scol).value if scol else None
    is_sum = isinstance(f, str) and "SUM" in f.upper()
    is_ref = (isinstance(f, str) and any(t in f for t in REF)) or (isinstance(v, str) and any(t in v for t in REF))
    is_num = isinstance(v, (int, float))

    if is_sum:
        n_sub += 1
        continue
    if is_ref:
        cur_cli += 1; cur_ref += 1
        continue
    if is_num:
        cur_cli += 1
        continue
    # sin saldo y no es SUM -> encabezado de vendedora (nuevo bloque)
    flush()
    cur_name = str(a).strip()

flush()

print(f"Vendedoras (bloques) detectadas: {len(vendedoras)}")
print(f"Subtotales (=SUM) saltados: {n_sub}\n")
print("  #  Vendedora                                  Clientes  (#REF!)")
print("  -- ------------------------------------------ --------  -------")
tot_cli = tot_ref = 0
for i, (name, ncli, nref) in enumerate(vendedoras, 1):
    tot_cli += ncli; tot_ref += nref
    print(f"  {i:>2} {name[:42]:<42} {ncli:>8}  {nref:>6}")
print(f"\n  TOTAL clientes en vendedoras: {tot_cli}  | con #REF!: {tot_ref}")

with open("tools/vendedoras.csv", "w", newline="", encoding="utf-8-sig") as fh:
    w = csv.writer(fh)
    w.writerow(["vendedora", "n_clientes", "n_ref"])
    w.writerows(vendedoras)
print("\n  (Lista guardada en tools/vendedoras.csv — local, no se commitea.)")
