#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Extractor del Kardex (Excel) -> CSV revisable para la migración del CRM (Bloque 5).

Lee SOLO las hojas activas 2026 (las 2024/2025 son histórico que se archiva, spec §8.4).
Por cada cliente saca: nombre crudo (col A, texto libre), saldo inicial (col B), y el
SALDO FINAL (la columna con la fórmula `=+B+D-E+...`, detectada automáticamente).
Marca los #REF! (saldo perdido, ~7 clientes — NO se inventan, se confirman con Daniel/Kary).

Salida: tools/migracion-review.csv (LOCAL, gitignored — datos reales). A consola: solo
AGREGADOS (conteos/totales), nunca nombres ni saldos individuales (repo público, L-15).

Uso:  python tools/extraer-kardex.py
"""
import csv
import re
import sys
from openpyxl import load_workbook

XLSX = "NUEVO KARDEX KARY DEL 2026....xlsx"
ACTIVE_SHEETS = ["kardex KARY 2026", "KARDEX VENDEDORAS  2026"]  # ojo: 2 espacios en VENDEDORAS
OUT = "tools/migracion-review.csv"

REF_TOKENS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def detect_saldo_col(ws_formulas):
    """La columna de saldo final = la de la fórmula con más referencias de celda
    (=+B2+D2-E2+...). Se busca en filas de muestra y se toma la columna más frecuente."""
    from collections import Counter
    votes = Counter()
    for row in ws_formulas.iter_rows(min_row=2, max_row=min(40, ws_formulas.max_row)):
        best_col, best_refs = None, 0
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and ("+" in v or "-" in v):
                refs = len(re.findall(r"[A-Z]{1,3}\d+", v))
                if refs > best_refs:
                    best_refs, best_col = refs, cell.column
        if best_col and best_refs >= 4:
            votes[best_col] += 1
    return votes.most_common(1)[0][0] if votes else None


def clean_name(raw):
    """Sugerencia de nombre limpio: quita montos, fechas y notas embebidas.
    SOLO sugerencia — el nombre final lo valida Daniel en el review."""
    if not raw:
        return ""
    s = str(raw)
    s = re.sub(r"(?i)cumplea\W?os.*$", "", s)        # "... CUMPLEAÑOS 09/11"
    s = re.sub(r"\d[\d.,]*", "", s)                   # montos / números
    s = re.sub(r"(?i)\b(ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)\b.*$", "", s)
    s = re.sub(r"[-_/]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def main():
    try:
        wbF = load_workbook(XLSX, data_only=False, read_only=True)
        wbV = load_workbook(XLSX, data_only=True, read_only=True)
    except Exception as e:
        print(f"ERROR abriendo {XLSX!r}: {e}")
        sys.exit(1)

    rows_out = []
    summary = {}

    for sheet in ACTIVE_SHEETS:
        if sheet not in wbF.sheetnames:
            print(f"  [!] hoja no encontrada: {sheet!r}")
            continue
        wsF, wsV = wbF[sheet], wbV[sheet]
        saldo_col = detect_saldo_col(wsF)
        n_total = n_ok = n_ref = n_sin = 0
        suma = 0.0

        for r in range(2, wsV.max_row + 1):
            raw = wsV.cell(r, 1).value          # col A
            if raw is None or str(raw).strip() == "":
                continue
            saldo_ini = wsV.cell(r, 2).value     # col B
            saldo_fin_f = wsF.cell(r, saldo_col).value if saldo_col else None
            saldo_fin_v = wsV.cell(r, saldo_col).value if saldo_col else None

            flag = "ok"
            if isinstance(saldo_fin_f, str) and any(t in saldo_fin_f for t in REF_TOKENS):
                flag = "REF-ERROR"; n_ref += 1
            elif isinstance(saldo_fin_v, str) and any(t in saldo_fin_v for t in REF_TOKENS):
                flag = "REF-ERROR"; n_ref += 1
            elif not isinstance(saldo_fin_v, (int, float)):
                flag = "SIN-SALDO"; n_sin += 1
            else:
                n_ok += 1
                suma += float(saldo_fin_v)

            n_total += 1
            rows_out.append({
                "hoja": sheet,
                "fila": r,
                "nombre_crudo": str(raw),
                "nombre_sugerido": clean_name(raw),
                "saldo_inicial": saldo_ini if isinstance(saldo_ini, (int, float)) else "",
                "saldo_final": saldo_fin_v if isinstance(saldo_fin_v, (int, float)) else "",
                "flag": flag,
            })

        summary[sheet] = {
            "saldo_col": saldo_col, "total": n_total, "ok": n_ok,
            "ref_error": n_ref, "sin_saldo": n_sin, "suma_ok": suma,
        }

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["hoja", "fila", "nombre_crudo", "nombre_sugerido",
                                          "saldo_inicial", "saldo_final", "flag"])
        w.writeheader()
        w.writerows(rows_out)

    # Consola: SOLO agregados (sin datos individuales)
    print(f"\n=== RESUMEN MIGRACIÓN (agregado) -> {OUT} ===")
    for sheet, s in summary.items():
        col = s["saldo_col"]
        from openpyxl.utils import get_column_letter
        colL = get_column_letter(col) if col else "?"
        print(f"\n  Hoja: {sheet}")
        print(f"    Columna de saldo detectada : {colL} (#{col})")
        print(f"    Filas con cliente          : {s['total']}")
        print(f"    Con saldo numérico OK      : {s['ok']}")
        print(f"    Con #REF! (saldo perdido)  : {s['ref_error']}  <-- confirmar con Daniel/Kary")
        print(f"    Sin saldo / no numérico    : {s['sin_saldo']}  <-- posibles subtotales/encabezados")
        print(f"    Suma de saldos OK (COP)    : {s['suma_ok']:,.0f}")
    print("\n  (El CSV con el detalle por cliente queda LOCAL, no se commitea.)")


if __name__ == "__main__":
    main()
